#!/usr/bin/env python3
"""[DEPRECATED] Anchor Excel v2 — 5-sheet workbook
已废弃 (2026-08-07)：统一使用 gen_excel_skill.py（10-sheet 完整版）。
本文件保留供历史参考，sync_all.py 已改调 gen_excel_skill.py。
"""
"""Anchor Excel Pro v2 — 5-sheet workbook"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

DATA = Path(r"C:\Users\lenovo\Desktop\portfolio_data.json")
OUT = Path(r"C:\Users\lenovo\Desktop\portfolio_holdings.xlsx")
OUT_ANCHOR = Path(r"C:\Users\lenovo\Desktop\Anchor\06-看板数据\portfolio_holdings.xlsx")

with open(DATA, encoding="utf-8") as f:
    D = json.load(f)

T = {
    'primary': '1a1a2e', 'accent': '5b8def', 'green': '00d4a0',
    'amber': 'f0b030', 'red': 'ff4d6a', 'white': 'ffffff',
    'text': '2d3748', 'text2': '6b7280', 'border': 'e5e8ed'
}
F_TITLE = Font(name='Arial', size=18, bold=True, color=T['primary'])
F_H2 = Font(name='Arial', size=14, bold=True, color=T['primary'])
F_HDR = Font(name='Arial', size=9, bold=True, color='ffffff')
F_BODY = Font(name='Arial', size=10, color=T['text'])
F_MONO = Font(name='Consolas', size=10, color=T['text'])
F_GREEN = Font(name='Consolas', size=10, color=T['green'])
F_RED = Font(name='Consolas', size=10, color=T['red'])
F_AMBER = Font(name='Consolas', size=10, color=T['amber'])
FILL_HDR = PatternFill(start_color=T['primary'], end_color=T['primary'], fill_type='solid')
FILL_ALT = PatternFill(start_color='f0f1f5', end_color='f0f1f5', fill_type='solid')
FILL_GREEN = PatternFill(start_color='e6f7f2', end_color='e6f7f2', fill_type='solid')
FILL_RED = PatternFill(start_color='ffe8ea', end_color='ffe8ea', fill_type='solid')
FILL_AMBER = PatternFill(start_color='fff8e6', end_color='fff8e6', fill_type='solid')
THIN = Side(style='thin', color=T['border'])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')
MNY = '#,##0.00'
PCT = '0.00%'

H = [h for h in D.get('holdings_summary', []) if h.get('mv', 0) > 1]
S = D.get('stock_holdings', [])
total = D.get('total_assets', 0)
pnl = D.get('total_hold_pnl_est', 0)
yb = D.get('yuebao', 0)
mkt = D.get('market', {})
update = D.get('update_date', '')

# Classify
bedrock, core_gr, satellite, pending = [], [], [], []
for h in H:
    n, g, mv = h.get('name', ''), h.get('group', ''), h.get('mv', 0)
    if g == '待结算': pending.append(h); continue
    if '余额宝' in n: continue
    if g == '全局固收' or '黄金' in n or '债券' in n: bedrock.append(h)
    elif '纳斯达克' in n or '纳指' in n or '天弘通利' in n: core_gr.append(h)
    else: satellite.append(h)
for s in S:
    bedrock.append({'name': s.get('name', '红利ETF'), 'mv': s.get('mv', 0),
                    'pnl': s.get('pnl', 0), 'day_pnl': s.get('day_pnl', 0),
                    'cumul': 0, 'note': f"{s.get('shares',0)}股"})

def write_header(ws, r, cols):
    for i, lbl in enumerate(cols, 1):
        c = ws.cell(row=r, column=i, value=lbl)
        c.font = F_HDR; c.fill = FILL_HDR; c.alignment = AC
    ws.row_dimensions[r].height = 22

def write_row(ws, r, vals, is_alt=False, colors=None):
    colors = colors or {}
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = colors.get(i, F_MONO if i > 1 else F_BODY)
        c.alignment = AR if i > 1 else AL
        if is_alt: c.fill = FILL_ALT

wb = Workbook()

# ===== Sheet 1: 总览 =====
ws = wb.active
ws.title = "Anchor总览"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for c, w in [(2, 26), (3, 13), (4, 14), (5, 14), (6, 14), (7, 14), (8, 14)]:
    ws.column_dimensions[get_column_letter(c)].width = w

ws.merge_cells('B2:H2'); ws['B2'] = "Anchor Investment System"; ws['B2'].font = F_TITLE
ws.row_dimensions[2].height = 34
ws.merge_cells('B3:H3'); ws['B3'] = f"净值日期 {update} | 总资产 ¥{total:,.2f} | 持有盈亏 {pnl:+,.0f}"
ws['B3'].font = Font(name='Arial', size=10, italic=True, color=T['text2'])

r = 5
kpi = [("总资产", f"¥{total:,.0f}", T['accent']), ("持有盈亏", f"{pnl:+,.0f}", T['green'] if pnl>=0 else T['red']),
       ("上证", str(mkt.get('sh',{}).get('close','--')), T['amber']), ("科创50", str(mkt.get('kc',{}).get('close','--')), T['red'])]
for i,(lbl,val,color) in enumerate(kpi):
    col = 2 + i*2
    ws.cell(row=r,column=col,value=lbl).font = Font(name='Arial',size=8,color=T['text2'],bold=True)
    cv = ws.cell(row=r+1,column=col,value=val); cv.font = Font(name='Arial',size=20,bold=True,color=color)
    if i<3: ws.merge_cells(f'{get_column_letter(col)}{r}:{get_column_letter(col+1)}{r}'); ws.merge_cells(f'{get_column_letter(col)}{r+1}:{get_column_letter(col+1)}{r+1}')

r = 9
layers = [("压舱石层 · 45%", bedrock, T['accent'], "永远不卖"), ("核心增长层 · 20%", core_gr, T['green'], "定投为主"),
          ("卫星进攻层 · 20%", satellite, T['amber'], "-8%止损"), ("现金预备层 · 15%", [], T['text2'], f"余额宝 ¥{yb:,.0f}")]
for lname, lfunds, color, rule in layers:
    lmv = sum(f.get('mv',0) for f in lfunds) if lfunds else yb
    ws.merge_cells(f'B{r}:H{r}')
    c = ws[f'B{r}']; c.value = f"  {lname} — ¥{lmv:,.0f} — {rule}"
    c.font = Font(name='Arial', size=12, bold=True, color=color)
    ws.row_dimensions[r].height = 24; r += 1
    if lfunds:
        write_header(ws, r, ["名称","市值","持有盈亏","累计盈亏","日收益","备注"])
        ws.row_dimensions[r].height = 22; r += 1
        for fi, f in enumerate(lfunds):
            vals = [f.get('name','?')[:40], f.get('mv',0), f.get('pnl',0) or 0, f.get('cumul',0) or 0, f.get('day_pnl',0) or 0, f.get('note','')[:25]]
            colors = {}
            for ci in (2,3,4):
                v = vals[ci]
                colors[ci+1] = F_GREEN if (v or 0) >= 0 else F_RED
            write_row(ws, r, vals, fi%2, colors)
            ws.cell(row=r,column=2).number_format = MNY
            for ci in (3,4,5): ws.cell(row=r,column=ci).number_format = MNY
            ws.row_dimensions[r].height = 20; r += 1
        r += 1
ws.freeze_panes = 'A2'

# ===== Sheet 2: 持仓明细 =====
ws2 = wb.create_sheet("持仓明细")
ws2.sheet_view.showGridLines = False
for c, w in [(1,3),(2,30),(3,12),(4,12),(5,10),(6,12),(7,12),(8,22)]:
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.merge_cells('B2:H2'); ws2['B2'] = f"全部持仓明细 — {update}"; ws2['B2'].font = F_H2
r = 4
write_header(ws2, r, ["基金名称","市值","持有盈亏","收益率","累计盈亏","日收益","备注/份额"]); r += 1
all_items = bedrock + core_gr + satellite + pending
all_items.sort(key=lambda x: x.get('mv',0), reverse=True)
for fi, f in enumerate(all_items):
    mv = f.get('mv',0); pv = f.get('pnl',0) or 0; cum = f.get('cumul',0) or 0; dp = f.get('day_pnl',0) or 0
    rate = pv/(mv-pv) if mv!=pv and mv>0 else 0
    vals = [f.get('name','?')[:38], mv, pv, rate, cum, dp, f.get('note','')[:30]]
    colors = {}
    for ci in (3,5,6):
        colors[ci] = F_GREEN if (vals[ci-1] or 0) >= 0 else F_RED
    colors[4] = F_GREEN if rate >= 0 else F_RED
    write_row(ws2, r, vals, fi%2, colors)
    ws2.cell(row=r,column=2).number_format = MNY
    ws2.cell(row=r,column=3).number_format = MNY
    ws2.cell(row=r,column=4).number_format = PCT
    ws2.cell(row=r,column=5).number_format = MNY
    ws2.cell(row=r,column=6).number_format = MNY
    ws2.row_dimensions[r].height = 20; r += 1
ws2.freeze_panes = 'B5'

# ===== Sheet 3: 规则检查 =====
ws3 = wb.create_sheet("规则检查")
ws3.sheet_view.showGridLines = False
for c, w in [(1,3),(2,28),(3,14),(4,14),(5,12),(6,20),(7,16)]:
    ws3.column_dimensions[get_column_letter(c)].width = w
ws3.merge_cells('B2:G2'); ws3['B2'] = "Anchor 规则检查 — 卫星层"; ws3['B2'].font = F_H2
r = 4
write_header(ws3, r, ["基金","市值","持有盈亏","收益率","状态","规则判断","建议"]); r += 1
for f in satellite:
    mv = f.get('mv',0); pv = f.get('pnl',0) or 0
    rate = pv/(mv-pv) if mv!=pv and mv>0 else 0
    if rate < -8: status, judge, suggest, fill = "止损", "触发-8%止损线", "反弹日清仓", FILL_RED
    elif rate < 0: status, judge, suggest, fill = "浮亏", "浮亏不加仓", "等待企稳", FILL_AMBER
    elif rate >= 10: status, judge, suggest, fill = "盈利", "可阶梯止盈", "+10%卖1/3", FILL_GREEN
    else: status, judge, suggest, fill = "持有", "健康", "持有", None
    vals = [f.get('name','?')[:26], mv, pv, rate, status, judge, suggest]
    for ci, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=ci, value=v)
        c.font = F_MONO if ci > 1 else F_BODY
        c.alignment = AR if ci > 1 else AL
        if fill: c.fill = fill
        if ci == 2: c.number_format = MNY
        if ci in (3,4): c.number_format = PCT if ci==4 else MNY; c.font = F_GREEN if (v or 0)>=0 else F_RED
    ws3.row_dimensions[r].height = 20; r += 1
ws3.freeze_panes = 'B5'

# ===== Sheet 4: 关注清单 & 待办 =====
ws4 = wb.create_sheet("关注清单")
ws4.sheet_view.showGridLines = False
for c, w in [(1,3),(2,14),(3,14),(4,16),(5,14),(6,12),(7,10),(8,14)]:
    ws4.column_dimensions[get_column_letter(c)].width = w
ws4.merge_cells('B2:H2'); ws4['B2'] = "关注清单 & 待办"; ws4['B2'].font = F_H2
r = 4
wl = D.get('watchlist', [])
write_header(ws4, r, ["排名","板块","ETF代码","触发条件","首笔金额","状态","",""]); r += 1
for w in wl:
    vals = [w.get('rank',''), w.get('sector',''), w.get('etf_code',w.get('etf','')), w.get('trigger',''), w.get('amount',''), w.get('status','')]
    for ci, v in enumerate(vals, 1):
        c = ws4.cell(row=r, column=ci, value=v); c.font = F_BODY; c.alignment = AC if ci==1 else AL
    ws4.row_dimensions[r].height = 20; r += 1
r += 1
ws4.merge_cells(f'B{r}:H{r}'); ws4[f'B{r}'] = "待办"; ws4[f'B{r}'].font = F_H2; r += 1
pa = D.get('pending_actions', [])
write_header(ws4, r, ["优先级","事项","","","","","",""]); r += 1
for p in pa:
    if isinstance(p, dict):
        vals = [p.get('priority',''), f"{p.get('action','')} · {p.get('name','')}", "", "", "", "", "", ""]
    else:
        vals = ["", p, "", "", "", "", "", ""]
    for ci, v in enumerate(vals, 1):
        c = ws4.cell(row=r, column=ci, value=v); c.font = F_BODY; c.alignment = AL
    ws4.row_dimensions[r].height = 20; r += 1
ws4.freeze_panes = 'B5'

# ===== Sheet 5: 交易记录 =====
ws5 = wb.create_sheet("交易记录")
ws5.sheet_view.showGridLines = False
for c, w in [(1,3),(2,14),(3,30),(4,14),(5,12),(6,32)]:
    ws5.column_dimensions[get_column_letter(c)].width = w
txs = D.get('transactions', [])
ws5.merge_cells('B2:F2'); ws5['B2'] = f"交易记录 · 共 {len(txs)} 条"; ws5['B2'].font = F_H2
r = 4
write_header(ws5, r, ["日期","基金名称","操作","金额","备注"]); r += 1
for tx in reversed(txs):
    op = tx.get('op',''); is_sell = any(k in str(op) for k in ('清仓','减仓','卖出'))
    vals = [tx.get('date',''), tx.get('name','')[:38], op, tx.get('amount',0), tx.get('note','')[:40]]
    colors = {3: F_RED if is_sell else F_GREEN}
    write_row(ws5, r, vals, False, colors)
    ws5.cell(row=r,column=4).number_format = MNY
    ws5.row_dimensions[r].height = 20; r += 1
ws5.freeze_panes = 'B5'

wb.save(OUT)
import shutil
shutil.copy2(OUT, OUT_ANCHOR)
print(f"Anchor Excel Pro v2 — {len(all_items)} funds · {len(txs)} transactions · 5 sheets")
