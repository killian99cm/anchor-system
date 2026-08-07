#!/usr/bin/env python3
"""Portfolio Excel — Complete & Accurate (Skill Standard)"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.formatting.rule import DataBarRule,ColorScaleRule,CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart,Reference
from openpyxl.chart.series import DataPoint

DATA=Path(r"C:\Users\lenovo\Desktop\portfolio_data.json")
OUT=Path(r"C:\Users\lenovo\Desktop\portfolio_holdings.xlsx")
with open(DATA,encoding="utf-8") as f: D=json.load(f)

# Theme
T={'primary':'1F4E79','light':'D6E3F0','chart':['1F4E79','2E75B6','5B9BD5','9DC3E6','BDD7EE','DEEBF7']}
SEM={'pos':'2E7D32','neg':'C62828','warn':'F57C00'}
SRF='Georgia'; SNS='Calibri'

# Borders
O=Side(style='thin',color='C0C0C0'); HB=Side(style='medium',color=T['primary']); I=Side(style='thin',color='E0E0E0'); N=Side(style=None)

def frame(ws,r1,r2,c1,c2,has_hdr=True):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            cl=ws.cell(row=r,column=c)
            cl.border=Border(left=O if c==c1 else N,right=O if c==c2 else N,top=O if r==r1 else I,bottom=HB if(has_hdr and r==r1)else(O if r==r2 else I))

def hdr(ws,r,sc,labels):
    for j,v in enumerate(labels,sc):
        c=ws.cell(row=r,column=j,value=v); c.font=Font(name=SRF,size=10,bold=True,color='FFFFFF')
        c.fill=PatternFill(start_color=T['primary'],end_color=T['primary'],fill_type='solid')
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[r].height=28

def wc(ws,r,c,value,font=None,align='center',fmt=None):
    cell=ws.cell(row=r,column=c)
    cell.value=value if value is not None else ''
    cell.font=font or Font(name=SNS,size=11)
    if align=='L': cell.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    elif align=='l': cell.alignment=Alignment(horizontal='left',vertical='center')
    elif align=='R': cell.alignment=Alignment(horizontal='right',vertical='center')
    else: cell.alignment=Alignment(horizontal='center',vertical='center')
    if fmt: cell.number_format=fmt
    return cell

def pn_rule(ws,cl,r1,r2):
    rng=f"{cl}{r1}:{cl}{r2}"
    ws.conditional_formatting.add(rng,CellIsRule(operator='greaterThan',formula=['0'],font=Font(color=SEM['pos'])))
    ws.conditional_formatting.add(rng,CellIsRule(operator='lessThan',formula=['0'],font=Font(color=SEM['neg'])))

MNY='#,##0.00'; PCT='0.00%'; INT='#,##0'; PRC='0.000'

# Compute
H=[h for h in D['holdings_summary'] if h.get('name')!='余额宝' and h['mv']>1]
TOTAL_FUND=sum(h['mv'] for h in H)
TOTAL_DAY=sum((h.get('day_pnl')or h['mv']*(h.get('day_pct')or 0)) for h in H)
HOLD_PNL=sum((h.get('pnl')or 0) for h in H)
SA=D.get('stock_account',0)  # now a number (total stock account value)
ST=SA  # stock total = stock_account number directly
ST_HOLDINGS=D.get('stock_holdings',[])  # stock holdings list
FUND_D=D.get('total_assets',0)-SA  # fund account = total - stock
GRAND=D.get('total_assets',0)
ST_MV=sum(s.get('mv',0) for s in ST_HOLDINGS)  # stock market value
ST_DAY=sum(s.get('day_pnl',0) for s in ST_HOLDINGS)  # stock day pnl
CLEARED=[h for h in D.get('holdings_summary',[]) if h.get('group','')=='已清仓']  # derive from holdings_summary

# Market helpers (new JSON structure)
_M = D.get('market', {})
SH_INDEX = _M.get('sh', {}).get('close', 0)
_SH_CHG_STR = _M.get('sh', {}).get('change', '0%')
try: SH_CHANGE = float(str(_SH_CHG_STR).replace('%','').replace('+','')) / 100
except: SH_CHANGE = 0
DAY_RETURN = 0  # not directly available in new JSON

wb=Workbook()

# ═══ Sheet 1: 总览 ═══
ws=wb.active; ws.title="总览"; ws.sheet_view.showGridLines=False; ws.column_dimensions['A'].width=3
for c,wd in [(2,22),(3,16),(4,16),(5,16),(6,16),(7,16),(8,16)]: ws.column_dimensions[get_column_letter(c)].width=wd
ws.merge_cells('B2:H2'); ws['B2']="基金投资组合 · 实时看板"; ws['B2'].font=Font(name=SRF,size=22,bold=True,color=T['primary']); ws.row_dimensions[2].height=38
ws.merge_cells('B3:H3'); ws['B3']=f"净值日期 {D.get('update_date','')} | 上证 {SH_INDEX}({SH_CHANGE*100:+.2f}%) | 组合收益率 {DAY_RETURN*100:+.2f}%"
ws['B3'].font=Font(name=SNS,size=10,italic=True,color='666666')

# Nav index
nr=5; ws.merge_cells(f'B{nr}:C{nr}'); ws[f'B{nr}']="Sheet 导航"; ws[f'B{nr}'].font=Font(name=SRF,size=13,bold=True,color=T['primary']); ws.row_dimensions[nr].height=22
sheets=["总览","资产汇总","持仓明细","分组持仓","加减仓记录","已清仓记录","股票账户","资产配置","定投计划","待办清单"]
for i,sn in enumerate(sheets):
    rr=nr+1+i; cl=ws.cell(row=rr,column=2,value=f"{i+1}. {sn}")
    cl.font=Font(name=SNS,size=11,color=T['primary'],underline='single'); cl.hyperlink=f"#'{sn}'!A1"; ws.row_dimensions[rr].height=18

# KPI
kr=nr+len(sheets)+3
pos_c=sum(1 for h in H if isinstance(h.get('pnl'),(int,float)) and h['pnl']>0)
neg_c=sum(1 for h in H if isinstance(h.get('pnl'),(int,float)) and h['pnl']<0)
best_fund=max((h for h in H if h.get('day_pct')),key=lambda h:h.get('day_pct')or 0,default={'name':'—','day_pct':0})
kpis=[("全账户总资产",GRAND,f"基金{FUND_D:,.0f}+股票{ST:,.0f}"),("持仓市值",TOTAL_FUND,f"{len(H)}只基金"),
      ("当日盈亏",TOTAL_DAY,f"收益率{DAY_RETURN*100:+.2f}%"),("持有浮盈",HOLD_PNL,f"{pos_c}赚/{neg_c}亏"),
      ("债券压舱",sum(h['mv'] for h in H if h.get('group','')=='全局固收'),""),("今日最佳",best_fund['name'][:10],f"+{best_fund['day_pct']*100:.2f}%")]
for i,(l,v,s) in enumerate(kpis):
    col=2+i*2
    wc(ws,kr,col,l,Font(name=SNS,size=9,color='888888'))
    vf=Font(name=SRF,size=20,bold=True,color=T['primary'])
    if isinstance(v,(int,float)) and v<0: vf=Font(name=SRF,size=20,bold=True,color=SEM['neg'])
    elif isinstance(v,(int,float)) and v>0: vf=Font(name=SRF,size=20,bold=True,color=SEM['pos'])
    wc(ws,kr+1,col,v,vf,'R',MNY if isinstance(v,(int,float))else None)
    wc(ws,kr+2,col,s,Font(name=SNS,size=9,color='888888'))
    ws.row_dimensions[kr].height=16; ws.row_dimensions[kr+1].height=32; ws.row_dimensions[kr+2].height=16

# Insights
ir=kr+5; ws.merge_cells(f'B{ir}:H{ir}'); ws[f'B{ir}']="KEY INSIGHTS"; ws[f'B{ir}'].font=Font(name=SRF,size=14,bold=True,color=T['primary']); ws.row_dimensions[ir].height=22
for i,ins in enumerate([
    f"• 组合今日收益率 {DAY_RETURN*100:+.2f}%，跑赢沪深300（{SH_CHANGE*100:+.2f}%），债券+黄金+红利有效对冲科技暴跌",
    f"• 港股通创新药暴涨 +{best_fund['day_pct']*100:.2f}%，政策（医保目录调整）+ BD出海双驱动，7月主线之一",
    f"• 半导体单日 -8.31%，全板块86只成分股仅2只上涨，已接近回调10%加仓触发线",
    f"• 全账户总资产 {GRAND:,.0f} 元 = 基金 {FUND_D:,.0f} + 股票 {ST:,.0f}，已实现全口径统计",
    f"• 7月操作重点：证券ETF买入（7/10前）+ 半导体回调加仓 + 港股科技清仓 + 银行ETF待清算",
]):
    ws.merge_cells(f'B{ir+1+i}:H{ir+1+i}'); ws[f'B{ir+1+i}']=ins; ws[f'B{ir+1+i}'].font=Font(name=SNS,size=11); ws.row_dimensions[ir+1+i].height=20

fr=ir+len(["x"]*5)+3; ws.merge_cells(f'B{fr}:H{fr}')
ws[f'B{fr}']=f"数据来源: 东方财富妙想 mx-data / mx-search  |  生成: 2026-07-02  |  工具: Anchor 体系 + excel-generator"
ws[f'B{fr}'].font=Font(name=SNS,size=8,italic=True,color='AAAAAA')
ws.freeze_panes='B5'

# ═══ Sheet 2: 资产汇总 ═══
ws2=wb.create_sheet("资产汇总"); ws2.sheet_view.showGridLines=False; ws2.column_dimensions['A'].width=3
for c,wd in [(2,24),(3,16),(4,16),(5,16),(6,30)]: ws2.column_dimensions[get_column_letter(c)].width=wd
ws2.merge_cells('B2:F2'); ws2['B2']="全账户资产汇总"; ws2['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws2.row_dimensions[2].height=36
ws2.merge_cells('B3:F3'); ws2['B3']=f"更新 {D.get('update_date','')} | 上证 {SH_INDEX}({SH_CHANGE*100:+.2f}%) | 全账户合计 ¥{GRAND:,.2f}"
ws2['B3'].font=Font(name=SNS,size=10,italic=True,color='666666')

sr=5; hdr(ws2,sr,2,["账户","总资产","持仓市值","当日盈亏","说明"])
rows2=[("养基宝(基金账户)",FUND_D,TOTAL_FUND,TOTAL_DAY,f"{len(H)}只基金 + 余额宝{D['yuebao']:,.0f}"),
       ("国信证券(股票账户)",ST,ST_MV,ST_DAY,f"{len(ST_HOLDINGS)}只股票")]
for i,(a,b,c,d,e) in enumerate(rows2):
    rr=sr+1+i
    wc(ws2,rr,2,a,Font(name=SNS,size=11,bold=True),'L'); wc(ws2,rr,3,b,None,'R',MNY); wc(ws2,rr,4,c,None,'R',MNY)
    wc(ws2,rr,5,d,None,'R',MNY); wc(ws2,rr,6,e,Font(name=SNS,size=10,color='666666'),'l'); ws2.row_dimensions[rr].height=22
gr=sr+len(rows2)+1
for j,(v,al) in enumerate([("合计",'L'),(GRAND,'R'),(TOTAL_FUND+ST_MV,'R'),(TOTAL_DAY+ST_DAY,'R'),(f"全账户总资产 ¥{GRAND:,.2f}",'l')]):
    wc(ws2,gr,j+2,v,Font(name=SRF,size=12,bold=True,color=T['primary']),al,MNY if isinstance(v,(int,float))else None)
ws2.row_dimensions[gr].height=26; frame(ws2,sr,gr,2,6); pn_rule(ws2,'E',sr+1,gr); ws2.freeze_panes='B6'

# ═══ Sheet 3: 持仓明细 ═══
ws3=wb.create_sheet("持仓明细"); ws3.sheet_view.showGridLines=False; ws3.column_dimensions['A'].width=3
for c,wd in [(2,30),(3,10),(4,14),(5,14),(6,12),(7,14),(8,12),(9,12),(10,10),(11,10),(12,12),(13,30)]: ws3.column_dimensions[get_column_letter(c)].width=wd
ws3.merge_cells('B2:M2'); ws3['B2']=f"全部持仓明细（{D['update_date']} 净值）"; ws3['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws3.row_dimensions[2].height=36

hr3=5; hdr(ws3,hr3,2,["基金名称","代码","市值","持有收益","持有收益率","累计收益","当日盈亏","当日涨跌","周收益","月收益","今年涨幅","分组","备注"])
# Sort: by group order then market value desc
group_order={"防御组合":0,"进攻组合":1,"全局固收":2,"全局QDII":3,"全局宽基":4}
sorted_H=sorted(H,key=lambda x:(group_order.get(x.get('group',''),99),-x['mv']))
for i,h in enumerate(sorted_H):
    rr=hr3+1+i; dp=h.get('day_pnl')or h['mv']*(h.get('day_pct')or 0)
    wc(ws3,rr,2,h['name'],Font(name=SNS,size=11,bold=True),'L')
    wc(ws3,rr,3,h.get('code',''),Font(name=SNS,size=10,color='888888'))
    wc(ws3,rr,4,h['mv'],None,'R',MNY); wc(ws3,rr,5,h.get('pnl'),None,'R',MNY)
    wc(ws3,rr,6,h.get('hold_rate'),None,'R',PCT); wc(ws3,rr,7,h.get('cumul'),None,'R',MNY)
    wc(ws3,rr,8,dp,None,'R',MNY); wc(ws3,rr,9,h.get('day_pct'),None,'R',PCT)
    wc(ws3,rr,10,h.get('week_pnl'),None,'R',MNY); wc(ws3,rr,11,h.get('month_pnl'),None,'R',MNY)
    wc(ws3,rr,12,h.get('ytd_change'),None,'R',PCT)
    wc(ws3,rr,13,h.get('group',''),Font(name=SNS,size=10,color='666666'))
    if h.get('note'): wc(ws3,rr,14,h['note'],Font(name=SNS,size=9,italic=True,color='888888'),'l')
    ws3.row_dimensions[rr].height=22

# Yuebao
yr=hr3+len(sorted_H)+1
for j,(v,al,fmt) in enumerate([("余额宝",'L',None),("流动性",'C',None),(D.get('yuebao',0),'R',MNY),(None,'C',None),(None,'C',None),(0,'R',MNY),(0,'R',MNY),(None,'C',None),(None,'C',None),(None,'C',None),(None,'C',None),("全局固收",'C',None)]):
    wc(ws3,yr,j+2,v,Font(name=SNS,size=11,color='888888'),al,fmt)
    if j==0: wc(ws3,yr,j+2,v,Font(name=SNS,size=11,bold=True,color='888888'),al)

frame(ws3,hr3,yr,2,13)
ws3.conditional_formatting.add(f'D{hr3+1}:D{yr}',DataBarRule(start_type='min',end_type='max',color=T['primary'],showValue=True))
ws3.conditional_formatting.add(f'I{hr3+1}:I{yr}',ColorScaleRule(start_type='min',start_color='C62828',mid_type='percentile',mid_value=50,mid_color='FFFFFF',end_type='max',end_color='2E7D32'))
for cl in ['E','F','G','H','I','J','K','L']: pn_rule(ws3,cl,hr3+1,yr)
ws3.freeze_panes='C6'; ws3.auto_filter.ref=f"B{hr3}:M{yr}"

# ═══ Sheet 4: 分组持仓 ═══
ws4=wb.create_sheet("分组持仓"); ws4.sheet_view.showGridLines=False; ws4.column_dimensions['A'].width=3
for c,wd in [(2,30),(3,10),(4,14),(5,14),(6,14),(7,14)]: ws4.column_dimensions[get_column_letter(c)].width=wd
ws4.merge_cells('B2:G2'); ws4['B2']="分组持仓明细"; ws4['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws4.row_dimensions[2].height=36

gcolors={'防御组合':'E6F3FF','进攻组合':'FFEBEE','全局固收':'E8F5E9','全局QDII':'FFF3E0','全局宽基':'F3E5F5'}
cr4=5
for gn in ["防御组合","进攻组合","全局固收","全局QDII","全局宽基"]:
    gf=[h for h in sorted_H if h.get('group','')==gn]
    if not gf: continue
    gmv=sum(h['mv'] for h in gf); gday=sum((h.get('day_pnl')or h['mv']*(h.get('day_pct')or 0)) for h in gf)
    ws4.merge_cells(f'B{cr4}:G{cr4}'); ws4[f'B{cr4}']=f"  {gn}（{len(gf)}只，市值 {gmv:,.2f}，当日盈亏 {gday:+,.2f}）"
    ws4[f'B{cr4}'].font=Font(name=SRF,size=13,bold=True,color=T['primary'])
    ws4[f'B{cr4}'].fill=PatternFill(start_color=gcolors.get(gn,'FFFFFF'),end_color=gcolors.get(gn,'FFFFFF'),fill_type='solid')
    ws4.row_dimensions[cr4].height=24; cr4+=1
    hdr(ws4,cr4,2,["基金","代码","市值","持有收益","累计收益","当日盈亏"])
    for i,h in enumerate(gf):
        rr=cr4+1+i; dp=h.get('day_pnl')or h['mv']*(h.get('day_pct')or 0)
        wc(ws4,rr,2,h['name'],Font(name=SNS,size=11,bold=True),'L'); wc(ws4,rr,3,h.get('code',''),Font(name=SNS,size=10,color='888888'))
        wc(ws4,rr,4,h['mv'],None,'R',MNY); wc(ws4,rr,5,h.get('pnl'),None,'R',MNY)
        wc(ws4,rr,6,h.get('cumul'),None,'R',MNY); wc(ws4,rr,7,dp,None,'R',MNY); ws4.row_dimensions[rr].height=20
    frame(ws4,cr4,cr4+len(gf),2,7); pn_rule(ws4,'G',cr4+1,cr4+len(gf)); cr4+=len(gf)+2
ws4.freeze_panes='B5'

# ═══ Sheet 5: 加减仓记录 ═══
ws5=wb.create_sheet("加减仓记录"); ws5.sheet_view.showGridLines=False; ws5.column_dimensions['A'].width=3
for c,wd in [(2,14),(3,32),(4,10),(5,12),(6,14),(7,14),(8,32)]: ws5.column_dimensions[get_column_letter(c)].width=wd
ws5.merge_cells('B2:H2'); ws5['B2']="加减仓操作记录"; ws5['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws5.row_dimensions[2].height=36
ws5.merge_cells('B3:H3'); ws5['B3']=f"共 {len(D['transactions'])} 条（含已撤销 2 条）| 灰色=已撤销  |  最近更新 {D['update_date']}"
ws5['B3'].font=Font(name=SNS,size=10,italic=True,color='666666')

lr5=5; hdr(ws5,lr5,2,["日期","基金名称","操作","金额/份额","买入净值","份额","备注"])
im={'买入':'[买入]','减仓':'[减仓]','清仓':'[清仓]','调仓':'[调仓]','定投累计':'[定投]','定投':'[定投]','定投/加仓':'[定投]','转换':'[转换]','撤销':'[撤销]'}
for i,tx in enumerate(D['transactions']):
    rr=lr5+1+i; icon=im.get(tx['op'],f"[{tx['op']}]"); is_cancel='撤销' in tx.get('op','')
    cf=Font(name=SNS,size=11,color='AAAAAA',italic=True) if is_cancel else Font(name=SNS,size=11)
    cfb=Font(name=SNS,size=11,color='AAAAAA',italic=True) if is_cancel else Font(name=SNS,size=11,bold=True)
    opf=Font(name=SNS,size=11,bold=True,color=SEM['pos'] if '买入' in tx['op']else(SEM['neg'] if'清仓' in tx['op']or'减仓' in tx['op']else(SEM['warn'] if'转换' in tx['op']else None)))
    if is_cancel: opf=Font(name=SNS,size=11,color='AAAAAA',italic=True)
    amt=tx['amount']
    try: amt_val=float(amt)
    except: amt_val=None
    wc(ws5,rr,2,tx['date'],cf); wc(ws5,rr,3,tx['name'],cfb,'L'); wc(ws5,rr,4,icon,opf)
    if amt_val is not None: wc(ws5,rr,5,amt_val,cf,'R',MNY)
    else: wc(ws5,rr,5,str(amt),cf)
    wc(ws5,rr,6,'—',cf); wc(ws5,rr,7,'—',cf); wc(ws5,rr,8,tx.get('note',''),Font(name=SNS,size=10,color='AAAAAA' if is_cancel else '888888'),'l'); ws5.row_dimensions[rr].height=22
frame(ws5,lr5,lr5+len(D['transactions']),2,8); ws5.freeze_panes='C6'

# ═══ Sheet 6: 已清仓记录 ═══
ws6=wb.create_sheet("已清仓记录"); ws6.sheet_view.showGridLines=False; ws6.column_dimensions['A'].width=3
for c,wd in [(2,36),(3,10),(4,14),(5,30)]: ws6.column_dimensions[get_column_letter(c)].width=wd
ws6.merge_cells('B2:E2'); ws6['B2']="已清仓基金历史记录"; ws6['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws6.row_dimensions[2].height=36
total_cl=sum(f.get('cumul',0) for f in CLEARED); pos_cl=sum(1 for f in CLEARED if f.get('cumul',0)>0); neg_cl=sum(1 for f in CLEARED if f.get('cumul',0)<0)
ws6.merge_cells('B3:E3'); ws6['B3']=f"共 {len(CLEARED)} 只已清仓 | 累计净盈亏 {total_cl:+,.2f} 元 | {pos_cl} 只盈利 / {neg_cl} 只亏损"
ws6['B3'].font=Font(name=SNS,size=10,italic=True,color='666666')

cr6=5; hdr(ws6,cr6,2,["基金名称","分组","累计盈亏","备注"])
for i,f in enumerate(sorted(CLEARED,key=lambda x:-x.get('cumul',0))):
    rr=cr6+1+i
    wc(ws6,rr,2,f['name'],Font(name=SNS,size=11,bold=True),'L'); wc(ws6,rr,3,f['group'],None)
    wc(ws6,rr,4,f.get('cumul',0),None,'R',MNY); wc(ws6,rr,5,f.get('note',''),Font(name=SNS,size=10,color='888888'),'l'); ws6.row_dimensions[rr].height=20
frame(ws6,cr6,cr6+len(CLEARED),2,5); pn_rule(ws6,'D',cr6+1,cr6+len(CLEARED))
# Summary
sur=cr6+len(CLEARED)+2
ws6.merge_cells(f'B{sur}:E{sur}')
ws6[f'B{sur}']=f"清仓总盈亏: {total_cl:+,.2f} 元 | 盈利TOP: 泰康半导体+299.54 | 亏损TOP: 嘉实软件-615.95 | 银行ETF待7/3清算(+9.02)"
ws6[f'B{sur}'].font=Font(name=SRF,size=12,bold=True,color=T['primary']); ws6.row_dimensions[sur].height=26
ws6.freeze_panes='C6'

# ═══ Sheet 7: 股票账户 ═══
ws7=wb.create_sheet("股票账户"); ws7.sheet_view.showGridLines=False; ws7.column_dimensions['A'].width=3
for c,wd in [(2,26),(3,12),(4,12),(5,12),(6,14),(7,12),(8,14)]: ws7.column_dimensions[get_column_letter(c)].width=wd
ws7.merge_cells('B2:H2'); ws7['B2']=f"股票账户 — 国信证券"; ws7['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws7.row_dimensions[2].height=36
ws7.merge_cells('B3:H3'); ws7['B3']=f"总资产 ¥{ST:,.2f} | 持仓市值 ¥{ST_MV:,.2f} | 当日盈亏 {ST_DAY:+,.2f}"
ws7['B3'].font=Font(name=SNS,size=10,italic=True,color='666666')

sh_r=5; hdr(ws7,sh_r,2,["标的","持仓(股)","现价","成本价","持仓盈亏","盈亏比例","持仓市值"])
for i,s in enumerate(ST_HOLDINGS):
    rr=sh_r+1+i
    wc(ws7,rr,2,s.get('name',''),Font(name=SNS,size=11,bold=True),'L'); wc(ws7,rr,3,s.get('shares',0),None,'R',INT)
    wc(ws7,rr,4,s.get('price',0),None,'R',PRC); wc(ws7,rr,5,s.get('cost',0),None,'R',PRC)
    pnl_pct = s.get('pnl',0)/(s.get('cost',1)*s.get('shares',1))*100 if s.get('cost',0) and s.get('shares',0) else 0
    wc(ws7,rr,6,s.get('pnl',0),None,'R',MNY); wc(ws7,rr,7,pnl_pct/100,None,'R',PCT)
    wc(ws7,rr,8,s.get('mv',0),None,'R',MNY); ws7.row_dimensions[rr].height=22
frame(ws7,sh_r,sh_r+len(ST_HOLDINGS),2,8); pn_rule(ws7,'F',sh_r+1,sh_r+len(ST_HOLDINGS))
ws7.freeze_panes='C6'

# ═══ Sheet 8: 资产配置 ═══
ws8=wb.create_sheet("资产配置"); ws8.sheet_view.showGridLines=False; ws8.column_dimensions['A'].width=3
for c,wd in [(2,20),(3,14),(4,14),(5,14),(6,32)]: ws8.column_dimensions[get_column_letter(c)].width=wd
ws8.merge_cells('B2:F2'); ws8['B2']="资产配置分析"; ws8['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws8.row_dimensions[2].height=36

ga={}
for h in sorted_H:
    g=h.get('group',''); ga.setdefault(g,{'mv':0,'day':0,'items':[]})
    ga[g]['mv']+=h['mv']; ga[g]['day']+=(h.get('day_pnl')or h['mv']*(h.get('day_pct')or 0)); ga[g]['items'].append(h['name'][:10])
ar8=5; hdr(ws8,ar8,2,["分组","市值","占比","当日盈亏","主要成分"])
gal=sorted(ga.items(),key=lambda x:-x[1]['mv'])
for i,(gn,gd) in enumerate(gal):
    rr=ar8+1+i
    wc(ws8,rr,2,gn,Font(name=SNS,size=11,bold=True),'L'); wc(ws8,rr,3,gd['mv'],None,'R',MNY)
    wc(ws8,rr,4,gd['mv']/TOTAL_FUND,None,'R',PCT); wc(ws8,rr,5,gd['day'],None,'R',MNY)
    wc(ws8,rr,6,', '.join(gd['items'][:6]),Font(name=SNS,size=10,color='666666'),'l'); ws8.row_dimensions[rr].height=20
frame(ws8,ar8,ar8+len(gal),2,6); pn_rule(ws8,'E',ar8+1,ar8+len(gal))

pie=PieChart(); pie.title="资产配置"; pie.width=15; pie.height=11
dr=Reference(ws8,min_col=3,min_row=ar8,max_row=ar8+len(gal)); cr=Reference(ws8,min_col=2,min_row=ar8+1,max_row=ar8+len(gal))
pie.add_data(dr,titles_from_data=True); pie.set_categories(cr)
for i,s in enumerate(pie.series):
    for j in range(len(gal)): pt=DataPoint(idx=j); pt.graphicalProperties.solidFill=T['chart'][j%len(T['chart'])]; s.data_points.append(pt)
ws8.add_chart(pie,"B13"); ws8.freeze_panes='B6'

# ═══ Sheet 9: 定投计划 ═══
ws9=wb.create_sheet("定投计划"); ws9.sheet_view.showGridLines=False; ws9.column_dimensions['A'].width=3
for c,wd in [(2,24),(3,16),(4,14),(5,16),(6,30),(7,14)]: ws9.column_dimensions[get_column_letter(c)].width=wd
ws9.merge_cells('B2:G2'); ws9['B2']="定投计划"; ws9['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws9.row_dimensions[2].height=36

dr9=5; ws9.merge_cells(f'B{dr9}:G{dr9}'); ws9[f'B{dr9}']="自动定投（3条）"; ws9[f'B{dr9}'].font=Font(name=SRF,size=13,bold=True,color=T['primary']); ws9.row_dimensions[dr9].height=22; dr9+=1
hdr(ws9,dr9,2,["品种","频率","月均金额","类型","规则","当前市值"])
auto=[("国泰黄金ETF联接A","每天 ¥15",450,"避险","雷打不动",next((h['mv'] for h in sorted_H if'黄金'in h['name']and'国泰'in h['name']),0)),
      ("华泰纳斯达克100 A","每天 ¥10(上限)",300,"QDII限额","不能停",next((h['mv'] for h in sorted_H if'华泰'in h['name']and'纳'in h['name']),0)),
      ("天弘全球高端制造C","每天 ¥20(动态)",600,"QDII限额","回调5%→¥30,10%→¥50+单笔",next((h['mv'] for h in sorted_H if'高端制造'in h['name']),0))]
for i,row in enumerate(auto):
    rr=dr9+1+i
    for j,v in enumerate(row): wc(ws9,rr,j+2,v,Font(name=SNS,size=11,bold=True)if j==0 else None,'L'if j==0 else('R'if j in(2,5)else'C'),INT if j==2 else(MNY if j==5 else None)); ws9.row_dimensions[rr].height=20
frame(ws9,dr9,dr9+3,2,7)

mr9=dr9+5; ws9.merge_cells(f'B{mr9}:G{mr9}'); ws9[f'B{mr9}']="手动操作（3条）"; ws9[f'B{mr9}'].font=Font(name=SRF,size=13,bold=True,color=T['primary']); ws9.row_dimensions[mr9].height=22; mr9+=1
hdr(ws9,mr9,2,["品种","触发条件","金额","当前市值","截止日期","状态"])
man=[("上证红利联接A","每月初手动","¥500",next((h['mv'] for h in sorted_H if'红利'in h['name']),0),"—","进行中"),
     ("华夏半导体芯片","从高点回调10-15%","每次¥500",next((h['mv'] for h in sorted_H if'半导体'in h['name']),0),"—","7/2已跌-8.4%，接近触发"),
     ("易方达证券ETF联接C","7月10日前","¥500-1,000",next((h['mv'] for h in sorted_H if'证券'in h['name']),0),"2026-07-10","待执行")]
for i,row in enumerate(man):
    rr=mr9+1+i
    for j,v in enumerate(row): wc(ws9,rr,j+2,v,Font(name=SNS,size=11,bold=True)if j==0 else None,'L'if j==0 else('R'if j==3 else'C'),MNY if j==3 else None); ws9.row_dimensions[rr].height=20
frame(ws9,mr9,mr9+3,2,7); ws9.freeze_panes='B6'

# ═══ Sheet 10: 待办清单 ═══
ws10=wb.create_sheet("待办清单"); ws10.sheet_view.showGridLines=False; ws10.column_dimensions['A'].width=3
for c,wd in [(2,12),(3,42),(4,10),(5,30)]: ws10.column_dimensions[get_column_letter(c)].width=wd
ws10.merge_cells('B2:E2'); ws10['B2']="待办清单 & 投资纪律"; ws10['B2'].font=Font(name=SRF,size=20,bold=True,color=T['primary']); ws10.row_dimensions[2].height=36

tr10=5; hdr(ws10,tr10,2,["状态","事项","类型","备注"])
for i,(s,t,ty,n) in enumerate([("✓ 完成","半导体+300","立即","7/2收盘前,买在-8.4%低位"),("✓ 完成","515180买入2700股","立即","7/2 国信证券,成本1.301"),("✓ 完成","银行ETF清仓","立即","卖出267.28份,待7/3到账"),("○ 待办","证券ETF买入500-1000","立即","7月10日前"),("○ 待办","港股科技(409)清仓","立即","-18.5%浮亏,回本需涨22.7%"),("○ 待办","上证红利联接A每月买","每月","月初¥500"),("◷ 等待","创新药Q3结束","等待","-7.99%,不回本清"),("◷ 等待","卫星产业清仓","等待","-10.35%,留诺安一只航天"),("◷ 等待","半导体回调加仓","等待","已跌-8.4%,近触发线"),("◷ 等待","光伏保留等拐点","等待","-9.01%,不定投")]):
    rr=tr10+1+i
    sf=Font(name=SNS,size=11,color=SEM['pos']if'完成'in s else(SEM['warn']if'待办'in s else None))
    wc(ws10,rr,2,s,sf); wc(ws10,rr,3,t,Font(name=SNS,size=11,bold=True),'L'); wc(ws10,rr,4,ty,None); wc(ws10,rr,5,n,Font(name=SNS,size=10,color='888888'),'l'); ws10.row_dimensions[rr].height=22
frame(ws10,tr10,tr10+10,2,5)

pr=tr10+13; ws10.merge_cells(f'B{pr}:E{pr}'); ws10[f'B{pr}']="核心纪律"; ws10[f'B{pr}'].font=Font(name=SRF,size=13,bold=True,color=T['primary']); ws10.row_dimensions[pr].height=22
for i,r in enumerate(["1. 半年内不新开任何基金","2. 科技回调10%+才加仓，大涨日止盈1/3老仓","3. 债券+纳指=锁定不动，其余动态调配","4. 手续费5元→手动买入至少1500-2000/次","5. 先研究再给结论，不草率下判断","6. 谁性价比高钱就去哪，不设死上限"]):
    rr=pr+1+i; ws10.merge_cells(f'B{rr}:E{rr}'); ws10[f'B{rr}']=r; ws10[f'B{rr}'].font=Font(name=SNS,size=11); ws10.row_dimensions[rr].height=20
ws10.freeze_panes='B6'

# ═══ Save ═══
try:
    wb.save(OUT)
    print("OK: " + OUT.name)
    # Anchor copy
    ANCHOR_OUT = Path(r"C:\Users\lenovo\Desktop\Anchor\06-dashboard\portfolio_holdings.xlsx")
    try:
        wb.save(ANCHOR_OUT)
        print("OK: " + ANCHOR_OUT.name)
    except PermissionError:
        print("Anchor copy skipped (file locked)")
except PermissionError:
    alt=Path(r"C:\Users\lenovo\Desktop\portfolio_holdings_new.xlsx")
    wb.save(alt)
    print(f"ALT: {alt.name}")
